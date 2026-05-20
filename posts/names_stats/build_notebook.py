#!/usr/bin/env python3
"""Generate czech_baby_names.ipynb — the full end-to-end notebook."""

import json, uuid
from pathlib import Path

# ── helpers ──────────────────────────────────────────────────────────────────

def _id():
    return uuid.uuid4().hex[:8]

def md(*lines):
    return {"cell_type": "markdown", "id": _id(), "metadata": {}, "source": list(lines)}

def code(*lines):
    return {
        "cell_type": "code", "id": _id(), "execution_count": None,
        "metadata": {}, "outputs": [],
        "source": list(lines),
    }

NL = "\n"

# ── cells ────────────────────────────────────────────────────────────────────

cells = []

# ── 0. title ─────────────────────────────────────────────────────────────────
cells += [md(
    "# Nejoblíbenější dětská jména v ČR\n",
    "**End-to-end pipeline:** discover → download → parse → visualise\n\n",
    "Data source: [Český statistický úřad (ČSÚ)](https://csu.gov.cz) — press releases on baby-name popularity, ",
    "covering years 2016–2019 (cumulative living-child counts) and 2022–2024 (annual births).\n\n",
    "> **Note on methodology:** 2016–2019 files list *all* registered names for every living child in the Czech Republic ",
    "as of January of that year (thousands of names ranked by count). ",
    "The 2022–2024 files list births *in that calendar year* only (top 100 nationally, top 20 per region).",
)]

# ── 1. deps ───────────────────────────────────────────────────────────────────
cells += [
    md("## 1. Dependencies\n\n",
       "Install with `uv` (fast) or plain pip — only needed once."),
    code(
        "# Uncomment to install\n",
        "# import subprocess, sys\n",
        "# subprocess.run([sys.executable, '-m', 'pip', 'install', 'requests', 'beautifulsoup4', 'openpyxl', 'pandas'], check=True)\n",
    ),
    code(
        "import csv, json, re, os\n",
        "from pathlib import Path\n",
        "from collections import defaultdict\n",
        "import requests\n",
        "from bs4 import BeautifulSoup\n",
        "import openpyxl\n",
        "import pandas as pd\n",
        "from IPython.display import display, HTML, IFrame\n",
    ),
]

# ── 2. links ──────────────────────────────────────────────────────────────────
cells += [
    md("## 2. Source links\n\n",
       "Seven press-release pages from ČSÚ — one per published year."),
    code(
        "links_raw = Path('links.txt').read_text(encoding='utf-8')\n",
        "print(links_raw)\n",
    ),
    code(
        "PAGES = [\n",
        "    (2016, 'https://csu.gov.cz/produkty/nejoblibenejsi-detska-jmena-jsou-jakub-a-eliska'),\n",
        "    (2017, 'https://csu.gov.cz/produkty/rodice-detem-nejcasteji-davaji-jmena-jan-a-eliska'),\n",
        "    (2018, 'https://csu.gov.cz/produkty/nejoblibenejsi-detska-jmena-jsou-jakub-a-eliska-fu0hnyje70'),\n",
        "    (2019, 'https://csu.gov.cz/produkty/eliska-a-jakub-opet-dominuji'),\n",
        "    (2022, 'https://csu.gov.cz/produkty/popularite-detskych-jmen-loni-vevodili-jakub-s-eliskou'),\n",
        "    (2023, 'https://csu.gov.cz/produkty/jmena-jakub-a-eliska-byla-loni-opet-nejoblibenejsi'),\n",
        "    (2024, 'https://csu.gov.cz/produkty/detskym-jmenum-loni-opet-kralovali-jakub-a-eliska'),\n",
        "]\n",
    ),
]

# ── 3. discover xlsx links ─────────────────────────────────────────────────────
cells += [
    md("## 3. Discover XLSX download links\n\n",
       "Each page links to one or more `.xlsx` files. We scrape the href attributes."),
    code(
        "BASE = 'https://csu.gov.cz'\n",
        "\n",
        "def find_xlsx_links(url):\n",
        "    r = requests.get(url, timeout=30)\n",
        "    r.raise_for_status()\n",
        "    soup = BeautifulSoup(r.text, 'html.parser')\n",
        "    found = []\n",
        "    for a in soup.find_all('a', href=True):\n",
        "        href = a['href']\n",
        "        if '.xlsx' in href.lower():\n",
        "            if not href.startswith('http'):\n",
        "                href = BASE + href\n",
        "            # strip query string for the filename, keep full URL for download\n",
        "            fname = href.split('/')[-1].split('?')[0]\n",
        "            found.append((fname, href))\n",
        "    return found\n",
        "\n",
        "discovered = {}\n",
        "for year, url in PAGES:\n",
        "    links = find_xlsx_links(url)\n",
        "    discovered[year] = links\n",
        "    print(f'{year}: {[f for f,_ in links]}')\n",
    ),
]

# ── 4. download ────────────────────────────────────────────────────────────────
cells += [
    md("## 4. Download XLSX files\n\n",
       "Files are saved to `raw/`. Already-downloaded files are skipped."),
    code(
        "RAW = Path('raw')\n",
        "RAW.mkdir(exist_ok=True)\n",
        "\n",
        "# Map each (year, filename) to a local name that encodes year and type.\n",
        "# Earlier years have a single 'overall' file; 2022–2024 have 'republika' and 'kraje'.\n",
        "def local_name(year, fname):\n",
        "    f = fname.lower()\n",
        "    if 'kraje' in f:    return f'{year}_kraje.xlsx'\n",
        "    if 'republika' in f: return f'{year}_republika.xlsx'\n",
        "    return f'{year}_overall.xlsx'\n",
        "\n",
        "to_download = []\n",
        "for year, links in discovered.items():\n",
        "    for fname, url in links:\n",
        "        lname = local_name(year, fname)\n",
        "        to_download.append((lname, url))\n",
        "\n",
        "for lname, url in to_download:\n",
        "    dest = RAW / lname\n",
        "    if dest.exists():\n",
        "        print(f'  skip  {lname}')\n",
        "        continue\n",
        "    r = requests.get(url, timeout=60)\n",
        "    r.raise_for_status()\n",
        "    dest.write_bytes(r.content)\n",
        "    print(f'  saved {lname}  ({len(r.content)//1024} KB)')\n",
        "\n",
        "print('\\nFiles in raw/:')\n",
        "for p in sorted(RAW.iterdir()):\n",
        "    print(f'  {p.name:40s}  {p.stat().st_size//1024:>4} KB')\n",
    ),
]

# ── 5. inspect structures ──────────────────────────────────────────────────────
cells += [
    md("## 5. Inspect file structures\n\n",
       "The files fall into three distinct formats — we need to understand each before writing the parser."),
    md("### 5a. 2016–2019 — single file, Czech Republic only\n\n",
       "One sheet per file, boys and girls listed side-by-side in paired columns."),
    code(
        "for year in [2016, 2017, 2018, 2019]:\n",
        "    wb = openpyxl.load_workbook(RAW / f'{year}_overall.xlsx', data_only=True)\n",
        "    ws = wb.active\n",
        "    rows = [r for r in ws.iter_rows(values_only=True) if any(c is not None for c in r)]\n",
        "    print(f'--- {year}  sheets={wb.sheetnames}  rows={ws.max_row}  cols={ws.max_column}')\n",
        "    print(f'    header: {rows[0]}')\n",
        "    for r in rows[1:4]: print(f'    {r}')\n",
        "    print()\n",
    ),
    md("### 5b. 2022–2024 republika — two sheets (Chlapci / Dívky), top 100\n\n",
       "Simple two-column layout: rank + name."),
    code(
        "for year in [2022, 2023, 2024]:\n",
        "    wb = openpyxl.load_workbook(RAW / f'{year}_republika.xlsx', data_only=True)\n",
        "    for sn in wb.sheetnames:\n",
        "        ws = wb[sn]\n",
        "        rows = list(ws.iter_rows(min_row=1, max_row=4, values_only=True))\n",
        "        print(f'{year} [{sn}]  {ws.max_row-1} data rows')\n",
        "        for r in rows: print(f'    {r}')\n",
        "    print()\n",
    ),
    md("### 5c. 2022–2024 kraje — two sheets, 15 regions × 3 columns (rank, name, spacer)\n\n",
       "Row 1 = region names, Row 2 = column headers, Rows 3+ = data (top ~20 per region)."),
    code(
        "wb = openpyxl.load_workbook(RAW / '2022_kraje.xlsx', data_only=True)\n",
        "ws = wb['Chlapci']\n",
        "rows = list(ws.iter_rows(values_only=True))\n",
        "print('Header row (region names):')\n",
        "print(' | '.join(str(v) for v in rows[0] if v is not None))\n",
        "print(f'\\nData rows per region (Chlapci):')\n",
        "for i, region in enumerate(['Česko','Praha','Středočeský','Jihočeský','Plzeňský',\n",
        "                              'Karlovarský','Ústecký','Liberecký','Královéhradecký',\n",
        "                              'Pardubický','Vysočina','Jihomoravský','Olomoucký',\n",
        "                              'Zlínský','Moravskoslezský']):\n",
        "    col_n = i * 3 + 1\n",
        "    count = sum(1 for r in rows[2:] if col_n < len(r) and r[col_n] is not None)\n",
        "    print(f'  {region:<22} {count} names')\n",
    ),
]

# ── 6. parse ───────────────────────────────────────────────────────────────────
cells += [
    md("## 6. Parse into unified CSV\n\n",
       "All formats are normalised to a single schema: `year, region, gender, rank, name`.\n\n",
       "**Strategy:**\n",
       "- 2016–2019: detect column offsets from the header row; extract boys/girls pairs; region = `Česko`\n",
       "- 2022–2024 kraje: primary source — includes all 15 regions, top ~20 names each\n",
       "- 2022–2024 republika: supplement — adds ranks 21–100 for `Česko` that kraje doesn't cover"),
    code(
        "REGIONS_KRAJE = [\n",
        "    'Česko', 'Praha', 'Středočeský', 'Jihočeský', 'Plzeňský',\n",
        "    'Karlovarský', 'Ústecký', 'Liberecký', 'Královéhradecký', 'Pardubický',\n",
        "    'Vysočina', 'Jihomoravský', 'Olomoucký', 'Zlínský', 'Moravskoslezský',\n",
        "]\n",
        "\n",
        "def parse_rank(val):\n",
        "    if val is None: return None\n",
        "    return str(val).strip().replace('–', '-').replace('—', '-')\n",
        "\n",
        "records = []\n",
        "\n",
        "def parse_old_format(path, year):\n",
        "    wb = openpyxl.load_workbook(path, data_only=True)\n",
        "    ws = wb.active\n",
        "    rows = list(ws.iter_rows(values_only=True))\n",
        "    header = rows[0]\n",
        "    boy_rank_col = next(i for i, v in enumerate(header) if v == 'Pořadí')\n",
        "    boy_name_col = boy_rank_col + 1\n",
        "    girl_rank_col = next(i for i, v in enumerate(header) if v == 'Pořadí' and i > boy_rank_col)\n",
        "    girl_name_col = girl_rank_col + 1\n",
        "    for row in rows[1:]:\n",
        "        b_rank, b_name = row[boy_rank_col], row[boy_name_col]\n",
        "        g_rank, g_name = row[girl_rank_col], row[girl_name_col]\n",
        "        if b_name: records.append(dict(year=year, region='Česko', gender='boy',  rank=parse_rank(b_rank), name=str(b_name).strip()))\n",
        "        if g_name: records.append(dict(year=year, region='Česko', gender='girl', rank=parse_rank(g_rank), name=str(g_name).strip()))\n",
        "\n",
        "def parse_republika(path, year):\n",
        "    wb = openpyxl.load_workbook(path, data_only=True)\n",
        "    for sn, gender in [('Chlapci', 'boy'), ('Dívky', 'girl')]:\n",
        "        for row in list(wb[sn].iter_rows(values_only=True))[1:]:\n",
        "            rank, name = row[0], row[1]\n",
        "            if name: records.append(dict(year=year, region='Česko', gender=gender, rank=parse_rank(rank), name=str(name).strip().title()))\n",
        "\n",
        "def parse_kraje(path, year):\n",
        "    wb = openpyxl.load_workbook(path, data_only=True)\n",
        "    for sn, gender in [('Chlapci', 'boy'), ('Dívky', 'girl')]:\n",
        "        rows = list(wb[sn].iter_rows(values_only=True))\n",
        "        for ri, region in enumerate(REGIONS_KRAJE):\n",
        "            rank_col, name_col = ri * 3, ri * 3 + 1\n",
        "            for row in rows[2:]:\n",
        "                rank = row[rank_col] if rank_col < len(row) else None\n",
        "                name = row[name_col] if name_col < len(row) else None\n",
        "                if name: records.append(dict(year=year, region=region, gender=gender, rank=parse_rank(rank), name=str(name).strip().title()))\n",
        "\n",
        "# --- run all parsers ---\n",
        "for year in [2016, 2017, 2018, 2019]:\n",
        "    parse_old_format(RAW / f'{year}_overall.xlsx', year)\n",
        "\n",
        "for year in [2022, 2023, 2024]:\n",
        "    parse_kraje(RAW / f'{year}_kraje.xlsx', year)\n",
        "    # add republika ranks beyond what kraje covers (~21-100)\n",
        "    wb_rep = openpyxl.load_workbook(RAW / f'{year}_republika.xlsx', data_only=True)\n",
        "    kraje_max = {}\n",
        "    for r in records:\n",
        "        if r['year'] == year and r['region'] == 'Česko':\n",
        "            try:\n",
        "                rk = int(r['rank'].split('-')[0])\n",
        "                kraje_max[r['gender']] = max(kraje_max.get(r['gender'], 0), rk)\n",
        "            except (ValueError, AttributeError): pass\n",
        "    for sn, gender in [('Chlapci', 'boy'), ('Dívky', 'girl')]:\n",
        "        threshold = kraje_max.get(gender, 0)\n",
        "        for row in list(wb_rep[sn].iter_rows(values_only=True))[1:]:\n",
        "            rank, name = row[0], row[1]\n",
        "            if name and isinstance(rank, int) and rank > threshold:\n",
        "                records.append(dict(year=year, region='Česko', gender=gender, rank=parse_rank(rank), name=str(name).strip().title()))\n",
        "\n",
        "# write CSV\n",
        "OUT = Path('names.csv')\n",
        "with open(OUT, 'w', newline='', encoding='utf-8') as f:\n",
        "    writer = csv.DictWriter(f, fieldnames=['year','region','gender','rank','name'])\n",
        "    writer.writeheader()\n",
        "    writer.writerows(records)\n",
        "\n",
        "print(f'Wrote {len(records):,} records → {OUT}')\n",
    ),
]

# ── 7. data preview ────────────────────────────────────────────────────────────
cells += [
    md("## 7. Data preview\n\n",
       "Quick look at the unified dataset."),
    code(
        "df = pd.read_csv('names.csv')\n",
        "print(df.shape)\n",
        "df.head(10)\n",
    ),
    code(
        "print('Years covered:', sorted(df.year.unique()))\n",
        "print('Regions:', sorted(df.region.unique()))\n",
        "print()\n",
        "print('Records per year × gender (Česko only):')\n",
        "print(df[df.region=='Česko'].groupby(['year','gender']).size().unstack())\n",
    ),
    code(
        "print('Top 5 boys by year (Česko):')\n",
        "top5 = (\n",
        "    df[(df.region=='Česko') & (df.gender=='boy')]\n",
        "    .assign(rank_int=lambda d: pd.to_numeric(d['rank'].str.split('-').str[0], errors='coerce'))\n",
        "    .query('rank_int <= 5')\n",
        "    .sort_values(['year','rank_int'])\n",
        "    .groupby('year')['name'].apply(list)\n",
        ")\n",
        "for yr, names in top5.items():\n",
        "    print(f'  {yr}: {names}')\n",
    ),
]

# ── 8. generate HTML ───────────────────────────────────────────────────────────
cells += [
    md("## 8. Generate the interactive HTML visualisation\n\n",
       "A self-contained bump chart in plain HTML + vanilla JS:\n",
       "- sticky single-line toolbar (gender, region, top-N, search)\n",
       "- Bézier connectors between years, dashed across the 2020–21 gap\n",
       "- hover → highlight one name's trajectory + show rank\n",
       "- search (≥ 3 chars) → highlight matching names with rank labels"),
]

# embed the full generate_html.py logic inline
GENERATE_HTML_SOURCE = r"""
YEARS_VIZ = [2016, 2017, 2018, 2019, 2022, 2023, 2024]

def read_data_for_viz():
    data = {}
    with open('names.csv', encoding='utf-8') as f:
        for r in csv.DictReader(f):
            y, reg, g = int(r['year']), r['region'], r['gender']
            try:
                rank = int(r['rank'].split('-')[0])
            except (ValueError, AttributeError):
                rank = 9999
            data.setdefault(y, {}).setdefault(reg, {'boy': [], 'girl': []})
            data[y][reg][g].append({'rank': rank, 'name': r['name']})
    for y in data:
        for reg in data[y]:
            for g in ('boy', 'girl'):
                data[y][reg][g].sort(key=lambda x: x['rank'])
    return data

data_viz = read_data_for_viz()
all_regions = sorted({r for y in data_viz for r in data_viz[y]})
regions = ['Česko'] + [r for r in all_regions if r != 'Česko']

out = {}
for y in YEARS_VIZ:
    out[y] = {}
    for reg in regions:
        e = data_viz.get(y, {}).get(reg, {'boy': [], 'girl': []})
        out[y][reg] = {'boy': e['boy'][:100], 'girl': e['girl'][:100]}

data_json   = json.dumps(out, ensure_ascii=False, separators=(',', ':'))
region_opts = '\n'.join(f'      <option value="{r}">{r}</option>' for r in regions)
"""

HTML_TEMPLATE_SOURCE = open('generate_html.py', 'r', encoding='utf-8').read()
# Extract just the HTML_TEMPLATE string content for the notebook cell
html_template_marker = "HTML_TEMPLATE = r\"\"\""

cells += [
    code(
        "# ── data → JSON ──────────────────────────────────────────────────────────\n",
        *[l + "\n" for l in GENERATE_HTML_SOURCE.strip().splitlines()],
    ),
]

# Read the actual HTML template from generate_html.py and use it
cells += [
    code(
        "# ── HTML template (truncated for display — full version in generate_html.py) ──\n",
        "# Read the template from generate_html.py and substitute placeholders\n",
        "import importlib.util, sys\n",
        "\n",
        "spec = importlib.util.spec_from_file_location('gen', 'generate_html.py')\n",
        "gen_mod = importlib.util.load_from_spec = None  # avoid running main\n",
        "\n",
        "# Simpler: just read the file and extract HTML_TEMPLATE\n",
        "src = Path('generate_html.py').read_text(encoding='utf-8')\n",
        "# The template is everything between HTML_TEMPLATE = r\"\"\" and the final \"\"\"\n",
        "match = re.search(r'HTML_TEMPLATE = r\"\"\"(.+?)\"\"\"', src, re.DOTALL)\n",
        "HTML_TEMPLATE = match.group(1)\n",
        "\n",
        "html = HTML_TEMPLATE.replace('%%DATA%%', data_json).replace('%%REGIONS%%', region_opts)\n",
        "Path('index.html').write_text(html, encoding='utf-8')\n",
        "print(f'Generated index.html  ({len(html)//1024} KB)')\n",
    ),
]

# ── 9. embed ───────────────────────────────────────────────────────────────────
cells += [
    md("## 9. Visualisation\n\n",
       "The chart is embedded below via `<iframe>`. If it appears blank, open the link directly — ",
       "some Jupyter environments restrict inline frames."),
    code(
        "link = Path('index.html').resolve().as_uri()\n",
        "display(HTML(\n",
        "    f'<p style=\"margin-bottom:8px\">'\n",
        "    f'<a href=\"{link}\" target=\"_blank\" '\n",
        "    f'style=\"font-size:14px;font-weight:600;color:#2563eb;text-decoration:none;'\n",
        "    f'padding:6px 12px;border:1px solid #2563eb;border-radius:6px;\">'\n",
        "    f'↗ Open index.html in new tab</a></p>'\n",
        "))\n",
        "\n",
        "# Embed directly — works in JupyterLab / classic Notebook when served from the same dir\n",
        "display(IFrame(src='index.html', width='100%', height=720))\n",
    ),
    md("---\n\n",
       "*Data: Český statistický úřad (ČSÚ) · Notebook generated automatically · ",
       "Visualisation: plain HTML + vanilla JS, no external dependencies.*"),
]

# ── assemble ──────────────────────────────────────────────────────────────────
notebook = {
    "nbformat": 4,
    "nbformat_minor": 5,
    "metadata": {
        "kernelspec": {
            "display_name": "Python 3",
            "language": "python",
            "name": "python3",
        },
        "language_info": {
            "name": "python",
            "version": "3.11.0",
        },
    },
    "cells": cells,
}

Path('czech_baby_names.ipynb').write_text(
    json.dumps(notebook, ensure_ascii=False, indent=1),
    encoding='utf-8',
)
print("Created czech_baby_names.ipynb")
