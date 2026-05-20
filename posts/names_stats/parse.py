"""
Parse Czech Statistical Authority children's name ranking files into a unified CSV.

File formats encountered:
  2016–2019 (single file/year): one sheet, boys & girls side by side
  2022–2024 republika: two sheets (Chlapci/Dívky), rank+name, Czech Republic only
  2022–2024 kraje: two sheets, 44 columns = 15 regions × (rank, name, spacer)
"""

import csv
import openpyxl
from pathlib import Path

RAW = Path("raw")
OUT = Path("names.csv")

REGIONS_KRAJE = [
    "Česko", "Praha", "Středočeský", "Jihočeský", "Plzeňský",
    "Karlovarský", "Ústecký", "Liberecký", "Královéhradecký", "Pardubický",
    "Vysočina", "Jihomoravský", "Olomoucký", "Zlínský", "Moravskoslezský",
]

records = []


def parse_rank(val):
    """Return the rank as a string; normalise dash variants."""
    if val is None:
        return None
    s = str(val).strip().replace("–", "-").replace("—", "-")
    return s


def parse_old_format(path: Path, year: int):
    """2016–2019: one sheet, side-by-side boys/girls, Czech Republic only."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    # Detect column offsets from the header row
    header = rows[0]
    # Find where "Pořadí" first appears (boys rank col)
    boy_rank_col = next(i for i, v in enumerate(header) if v == "Pořadí")
    boy_name_col = boy_rank_col + 1
    # Girls: second "Pořadí"
    girl_rank_col = next(
        i for i, v in enumerate(header) if v == "Pořadí" and i > boy_rank_col
    )
    girl_name_col = girl_rank_col + 1

    for row in rows[1:]:
        b_rank, b_name = row[boy_rank_col], row[boy_name_col]
        g_rank, g_name = row[girl_rank_col], row[girl_name_col]
        if b_name:
            records.append(
                dict(year=year, region="Česko", gender="boy",
                     rank=parse_rank(b_rank), name=str(b_name).strip())
            )
        if g_name:
            records.append(
                dict(year=year, region="Česko", gender="girl",
                     rank=parse_rank(g_rank), name=str(g_name).strip())
            )


def parse_republika(path: Path, year: int):
    """2022–2024 republika: sheets Chlapci/Dívky, columns: rank, name."""
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name, gender in [("Chlapci", "boy"), ("Dívky", "girl")]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        for row in rows[1:]:  # skip header
            rank, name = row[0], row[1]
            if name:
                records.append(
                    dict(year=year, region="Česko", gender=gender,
                         rank=parse_rank(rank), name=str(name).strip().title())
                )


def parse_kraje(path: Path, year: int):
    """2022–2024 kraje: sheets Chlapci/Dívky, 15 regions × 3 columns each."""
    wb = openpyxl.load_workbook(path, data_only=True)
    for sheet_name, gender in [("Chlapci", "boy"), ("Dívky", "girl")]:
        ws = wb[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        # rows[0] = region names, rows[1] = column headers, rows[2+] = data
        for region_idx, region in enumerate(REGIONS_KRAJE):
            rank_col = region_idx * 3
            name_col = rank_col + 1
            for row in rows[2:]:
                rank = row[rank_col] if rank_col < len(row) else None
                name = row[name_col] if name_col < len(row) else None
                if name:
                    records.append(
                        dict(year=year, region=region, gender=gender,
                             rank=parse_rank(rank), name=str(name).strip().title())
                    )


# ── 2016–2019: single file, Czech Republic only ──────────────────────────────
for year in [2016, 2017, 2018, 2019]:
    parse_old_format(RAW / f"{year}_overall.xlsx", year)

# ── 2022–2024: separate republika + kraje files ───────────────────────────────
# Use kraje files as the primary source (they include Česko + all regions).
# The republika files have top-100 while kraje have top-20; we prefer kraje
# for consistency across years.  Add republika rows only when region == Česko
# and the rank exceeds what's in kraje.
for year in [2022, 2023, 2024]:
    parse_kraje(RAW / f"{year}_kraje.xlsx", year)
    # republika goes up to rank 100; kraje only ~20 for Česko — add the rest
    wb_rep = openpyxl.load_workbook(RAW / f"{year}_republika.xlsx", data_only=True)
    kraje_max = {}  # gender -> max numeric rank already in kraje for Česko
    for r in records:
        if r["year"] == year and r["region"] == "Česko":
            try:
                rk = int(r["rank"].split("-")[0])
                key = r["gender"]
                kraje_max[key] = max(kraje_max.get(key, 0), rk)
            except (ValueError, AttributeError):
                pass
    for sheet_name, gender in [("Chlapci", "boy"), ("Dívky", "girl")]:
        ws = wb_rep[sheet_name]
        rows = list(ws.iter_rows(values_only=True))
        threshold = kraje_max.get(gender, 0)
        for row in rows[1:]:
            rank, name = row[0], row[1]
            if name and isinstance(rank, int) and rank > threshold:
                records.append(
                    dict(year=year, region="Česko", gender=gender,
                         rank=parse_rank(rank), name=str(name).strip().title())
                )

# ── Write CSV ─────────────────────────────────────────────────────────────────
fields = ["year", "region", "gender", "rank", "name"]
with open(OUT, "w", newline="", encoding="utf-8") as f:
    writer = csv.DictWriter(f, fieldnames=fields)
    writer.writeheader()
    writer.writerows(records)

print(f"Wrote {len(records)} records to {OUT}")

# Quick sanity check
from collections import Counter
by_year_region = Counter((r["year"], r["region"]) for r in records)
print("\nRecords per year × region (sample):")
for (yr, region), count in sorted(by_year_region.items()):
    print(f"  {yr}  {region:<25} {count:>5}")
